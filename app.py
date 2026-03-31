"""
月报竞品新闻渠道统计分析工具 - Web服务
支持文件上传，自动处理并返回结果
"""

import os
import sys
import pandas as pd
from datetime import datetime
from flask import Flask, request, send_file, render_template_string, flash
from io import BytesIO

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'competitor-news-analyzer-secret-key')

# 媒体类别映射表（与命令行版本一致）
MEDIA_CATEGORIES = {
    # ========== 汽车行业垂类媒体 ==========
    "汽车之家": "汽车行业垂类媒体",
    "懂车帝": "汽车行业垂类媒体",
    "易车": "汽车行业垂类媒体",
    "太平洋汽车": "汽车行业垂类媒体",
    "爱卡汽车": "汽车行业垂类媒体",
    "网易汽车": "汽车行业垂类媒体",
    "新浪汽车": "汽车行业垂类媒体",
    "腾讯汽车": "汽车行业垂类媒体",
    "凤凰汽车": "汽车行业垂类媒体",
    "搜狐汽车": "汽车行业垂类媒体",
    "车质网": "汽车行业垂类媒体",
    "中国汽车报": "汽车行业垂类媒体",
    "autohome": "汽车行业垂类媒体",
    "dongchedi": "汽车行业垂类媒体",
    "yiche": "汽车行业垂类媒体",
    "Bil24": "汽车行业垂类媒体",
    "BilNorge.no": "汽车行业垂类媒体",
    "Bilnytt.no": "汽车行业垂类媒体",
    "Nybiltester": "汽车行业垂类媒体",
    "Nyheter fra Bilglede": "汽车行业垂类媒体",
    "Magasinet Motor": "汽车行业垂类媒体",
    "Elbil24.no": "汽车行业垂类媒体",
    "OtonomHaber (NO)": "汽车行业垂类媒体",

    # ========== 商业类媒体 ==========
    "第一财经": "商业类媒体",
    "每日经济新闻": "商业类媒体",
    "经济观察报": "商业类媒体",
    "36氪": "商业类媒体",
    "虎嗅": "商业类媒体",
    "钛媒体": "商业类媒体",
    "财经网": "商业类媒体",
    "华尔街见闻": "商业类媒体",
    "界面新闻": "商业类媒体",
    "雪球": "商业类媒体",
    "东方财富": "商业类媒体",
    "E24.no": "商业类媒体",
    "Finansavisen": "商业类媒体",
    "Offshore.no": "商业类媒体",
    "Teknisk Ukeblad Online": "商业类媒体",
    "ITavisen.no": "商业类媒体",
    "Lyd & Bilde Online": "商业类媒体",
    "Tek.no": "商业类媒体",
    "Gamereactor (NO)": "商业类媒体",
    "Teksiden": "商业类媒体",

    # ========== 综合类媒体 ==========
    "人民日报": "综合类媒体",
    "新华社": "综合类媒体",
    "央视": "综合类媒体",
    "人民网": "综合类媒体",
    "新华网": "综合类媒体",
    "澎湃新闻": "综合类媒体",
    "今日头条": "综合类媒体",
    "腾讯新闻": "综合类媒体",
    "网易新闻": "综合类媒体",
    "新浪新闻": "综合类媒体",
    "搜狐新闻": "综合类媒体",
    "凤凰资讯": "综合类媒体",
    "TV2 Nyheter": "综合类媒体",
    "Dagbladet Online": "综合类媒体",
    "VG Online": "综合类媒体",
    "Nordlys.no": "综合类媒体",
    "The News (NO)": "综合类媒体",
    "Norway News (NO)": "综合类媒体",
    "Vestlandsnytt": "综合类媒体",
    "Dagens (NO)": "综合类媒体",
    "MSN": "综合类媒体",

    # ========== 生活类媒体 ==========
    "微信": "生活类媒体",
    "微博": "生活类媒体",
    "小红书": "生活类媒体",
    "抖音": "生活类媒体",
    "快手": "生活类媒体",
    "Bilibili": "生活类媒体",
    "知乎": "生活类媒体",
    "百度": "生活类媒体",
    "UC": "生活类媒体",
    "Dinside.no": "生活类媒体",
    "Bodonu": "生活类媒体",
}


def format_audience(value):
    """格式化触达人数，添加单位，不四舍五入"""
    if pd.isna(value) or value == 0:
        return "-"
    if value >= 1_000_000:
        result = value / 1_000_000
        integer_part = int(result)
        decimal_part = int((result - integer_part) * 10)
        return f"{integer_part}.{decimal_part}M"
    elif value >= 1_000:
        result = value / 1_000
        integer_part = int(result)
        decimal_part = int((result - integer_part) * 10)
        return f"{integer_part}.{decimal_part}K"
    else:
        return str(int(value))


def get_media_category(media_name):
    """获取媒体类别"""
    if pd.isna(media_name):
        return "综合类媒体"
    media_name_str = str(media_name).strip()
    if media_name_str in MEDIA_CATEGORIES:
        return MEDIA_CATEGORIES[media_name_str]
    for key, category in MEDIA_CATEGORIES.items():
        if key in media_name_str or media_name_str in key:
            return category
    return "综合类媒体"


def analyze_competitor_news(df):
    """分析竞品新闻渠道数据"""
    # 检查必要的列
    required_columns = ["Media ID", "Date", "Source", "Potential Audience"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        return None, f"缺少必要的列: {', '.join(missing_columns)}"

    # 转换日期列
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # 获取唯一品牌
    brands = df["Media ID"].dropna().unique().tolist()

    # 存储所有结果
    all_results = []

    for brand in brands:
        brand_df = df[df["Media ID"] == brand].copy()

        # 统计每个媒体的声量（篇数）
        volume_df = brand_df.groupby("Source").size().reset_index(name="声量（篇）")

        # 获取每个媒体最新日期的触达数
        def get_latest_audience(media_name):
            media_data = brand_df[brand_df["Source"] == media_name]
            latest_row = media_data.loc[media_data["Date"].idxmax()] if not media_data.empty else None
            if latest_row is not None:
                return latest_row["Potential Audience"]
            return None

        volume_df["触达（人次）"] = volume_df["Source"].apply(get_latest_audience)

        # 按声量降序排列，取TOP10
        volume_df = volume_df.sort_values("声量（篇）", ascending=False).head(10)
        volume_df["品牌名称"] = brand
        volume_df["媒体类别"] = volume_df["Source"].apply(get_media_category)
        volume_df["触达（人次）"] = volume_df["触达（人次）"].apply(format_audience)
        volume_df.insert(0, "序号", range(1, len(volume_df) + 1))
        volume_df = volume_df.rename(columns={"Source": "媒体名称"})
        volume_df = volume_df[["序号", "品牌名称", "媒体名称", "声量（篇）", "触达（人次）", "媒体类别"]]

        all_results.append(volume_df)

    return all_results, None


# HTML模板
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>月报竞品新闻渠道统计分析</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 40px 20px;
        }
        .container {
            max-width: 900px;
            margin: 0 auto;
        }
        .card {
            background: white;
            border-radius: 16px;
            padding: 40px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
        }
        h1 {
            color: #333;
            text-align: center;
            margin-bottom: 10px;
            font-size: 28px;
        }
        .subtitle {
            text-align: center;
            color: #666;
            margin-bottom: 30px;
            font-size: 14px;
        }
        .upload-area {
            border: 2px dashed #ddd;
            border-radius: 12px;
            padding: 40px;
            text-align: center;
            margin-bottom: 20px;
            transition: all 0.3s;
            cursor: pointer;
        }
        .upload-area:hover {
            border-color: #667eea;
            background: #f8f9ff;
        }
        .upload-area.dragover {
            border-color: #667eea;
            background: #f0f3ff;
        }
        .upload-icon {
            font-size: 48px;
            margin-bottom: 15px;
        }
        .upload-text {
            color: #666;
            font-size: 16px;
        }
        .upload-text span {
            color: #667eea;
            font-weight: 600;
        }
        input[type="file"] {
            display: none;
        }
        .btn {
            display: inline-block;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 14px 32px;
            border: none;
            border-radius: 8px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: transform 0.2s, box-shadow 0.2s;
        }
        .btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.4);
        }
        .btn:disabled {
            background: #ccc;
            cursor: not-allowed;
            transform: none;
            box-shadow: none;
        }
        .submit-area {
            text-align: center;
            margin-top: 20px;
        }
        .note {
            background: #f8f9fa;
            border-radius: 8px;
            padding: 16px;
            margin-top: 20px;
            font-size: 13px;
            color: #666;
        }
        .note h3 {
            color: #333;
            margin-bottom: 8px;
            font-size: 14px;
        }
        .note ul {
            margin-left: 20px;
        }
        .result-info {
            background: #d4edda;
            border-radius: 8px;
            padding: 16px;
            margin-bottom: 20px;
            color: #155724;
        }
        .brand-section {
            margin-bottom: 30px;
        }
        .brand-title {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 10px 20px;
            border-radius: 8px 8px 0 0;
            font-weight: 600;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }
        th {
            background: #f8f9fa;
            padding: 12px 8px;
            text-align: center;
            font-weight: 600;
            color: #333;
            border-bottom: 2px solid #dee2e6;
        }
        td {
            padding: 10px 8px;
            text-align: center;
            border-bottom: 1px solid #dee2e6;
        }
        tr:last-child td {
            border-bottom: none;
        }
        tr:hover {
            background: #f8f9fa;
        }
        .category-tag {
            display: inline-block;
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 500;
        }
        .category-汽车 { background: #e3f2fd; color: #1565c0; }
        .category-商业 { background: #fff3e0; color: #e65100; }
        .category-综合 { background: #f3e5f5; color: #7b1fa2; }
        .category-生活 { background: #e8f5e9; color: #2e7d32; }
        .error {
            background: #f8d7da;
            color: #721c24;
            padding: 16px;
            border-radius: 8px;
            margin-bottom: 20px;
        }
        @media (max-width: 600px) {
            .card { padding: 20px; }
            table { font-size: 12px; }
            th, td { padding: 8px 4px; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="card">
            <h1>月报竞品新闻渠道统计分析</h1>
            <p class="subtitle">上传Excel文件，自动统计品牌声量TOP10媒体</p>

            {% with messages = get_flashed_messages() %}
                {% if messages %}
                    {% for message in messages %}
                        <div class="error">{{ message }}</div>
                    {% endfor %}
                {% endif %}
            {% endwith %}

            <form method="POST" enctype="multipart/form-data" id="uploadForm">
                <div class="upload-area" id="uploadArea">
                    <div class="upload-icon">📁</div>
                    <p class="upload-text">
                        点击或拖拽文件到此处上传<br>
                        <span>支持 .xlsx 格式</span>
                    </p>
                    <input type="file" name="file" id="fileInput" accept=".xlsx">
                    <p id="fileName" style="margin-top: 10px; color: #667eea; font-weight: 500;"></p>
                </div>
                <div class="submit-area">
                    <button type="submit" class="btn" id="submitBtn">开始分析</button>
                </div>
            </form>

            {% if result %}
                <div class="result-info">
                    ✅ 分析完成！共处理 <strong>{{ brands_count }}</strong> 个品牌，生成 <strong>{{ total_media }}</strong> 条记录
                </div>

                {% for brand_df in result %}
                    <div class="brand-section">
                        <div class="brand-title">{{ brand_df[0]['品牌名称'] }} TOP10</div>
                        <table>
                            <thead>
                                <tr>
                                    <th>序号</th>
                                    <th>媒体名称</th>
                                    <th>声量（篇）</th>
                                    <th>触达（人次）</th>
                                    <th>媒体类别</th>
                                </tr>
                            </thead>
                            <tbody>
                                {% for _, row in brand_df.iterrows() %}
                                <tr>
                                    <td>{{ row['序号'] }}</td>
                                    <td>{{ row['媒体名称'] }}</td>
                                    <td>{{ row['声量（篇）'] }}</td>
                                    <td>{{ row['触达（人次）'] }}</td>
                                    <td>
                                        {% if '汽车' in row['媒体类别'] %}
                                            <span class="category-tag category-汽车">{{ row['媒体类别'] }}</span>
                                        {% elif '商业' in row['媒体类别'] %}
                                            <span class="category-tag category-商业">{{ row['媒体类别'] }}</span>
                                        {% elif '生活' in row['媒体类别'] %}
                                            <span class="category-tag category-生活">{{ row['媒体类别'] }}</span>
                                        {% else %}
                                            <span class="category-tag category-综合">{{ row['媒体类别'] }}</span>
                                        {% endif %}
                                    </td>
                                </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                {% endfor %}

                <div class="submit-area">
                    <a href="{{ download_url }}" class="btn">📥 下载结果文件</a>
                </div>
            {% endif %}

            <div class="note">
                <h3>📋 使用说明</h3>
                <ul>
                    <li>上传包含 <strong>Media ID</strong>、<strong>Date</strong>、<strong>Source</strong>、<strong>Potential Audience</strong> 列的Excel文件</li>
                    <li>Media ID 对应品牌名称，Source 对应媒体名称</li>
                    <li>触达（人次）以 Date 最晚的数据为准</li>
                    <li>结果按声量（篇）降序排列，每个品牌显示 TOP10</li>
                    <li>触达单位：K（千）、M（百万），数值保留一位小数，不四舍五入</li>
                </ul>
            </div>
        </div>
    </div>

    <script>
        const uploadArea = document.getElementById('uploadArea');
        const fileInput = document.getElementById('fileInput');
        const fileName = document.getElementById('fileName');
        const submitBtn = document.getElementById('submitBtn');

        uploadArea.addEventListener('click', () => fileInput.click());

        uploadArea.addEventListener('dragover', (e) => {
            e.preventDefault();
            uploadArea.classList.add('dragover');
        });

        uploadArea.addEventListener('dragleave', () => {
            uploadArea.classList.remove('dragover');
        });

        uploadArea.addEventListener('drop', (e) => {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
            if (e.dataTransfer.files.length) {
                fileInput.files = e.dataTransfer.files;
                fileName.textContent = e.dataTransfer.files[0].name;
            }
        });

        fileInput.addEventListener('change', () => {
            if (fileInput.files.length) {
                fileName.textContent = fileInput.files[0].name;
            }
        });

        document.getElementById('uploadForm').addEventListener('submit', () => {
            submitBtn.disabled = true;
            submitBtn.textContent = '处理中...';
        });
    </script>
</body>
</html>
'''


@app.route('/', methods=['GET', 'POST'])
def index():
    result = None
    error = None
    brands_count = 0
    total_media = 0
    download_url = None

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择文件')
            return render_template_string(HTML_TEMPLATE)

        file = request.files['file']
        if file.filename == '':
            flash('请选择文件')
            return render_template_string(HTML_TEMPLATE)

        if not file.filename.endswith('.xlsx'):
            flash('只支持 .xlsx 格式文件')
            return render_template_string(HTML_TEMPLATE)

        try:
            # 读取文件
            df = pd.read_excel(file)

            # 分析数据
            result, error = analyze_competitor_news(df)

            if error:
                flash(error)
                return render_template_string(HTML_TEMPLATE)

            # 生成结果文件
            output = BytesIO()
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "处理结果"

            header_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            row_num = 1
            for i, brand_df in enumerate(result):
                if i > 0:
                    row_num += 1
                headers = ["序号", "品牌名称", "媒体名称", "声量（篇）", "触达（人次）", "媒体类别"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col, value=header)
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                row_num += 1

                for _, row_data in brand_df.iterrows():
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col, value=value)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                    row_num += 1

            column_widths = [8, 12, 25, 12, 14, 18]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width

            wb.save(output)
            output.seek(0)

            # 保存到临时文件
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            temp_filename = f'result_{timestamp}.xlsx'
            temp_path = os.path.join('/tmp' if os.path.exists('/tmp') else '.', temp_filename)
            with open(temp_path, 'wb') as f:
                f.write(output.getvalue())
            download_url = f'/download/{temp_filename}'

            brands_count = len(result)
            total_media = sum(len(r) for r in result)

        except Exception as e:
            flash(f'处理失败: {str(e)}')
            return render_template_string(HTML_TEMPLATE)

    return render_template_string(HTML_TEMPLATE, result=result, error=error,
                                   brands_count=brands_count, total_media=total_media,
                                   download_url=download_url)


@app.route('/download/<filename>')
def download(filename):
    return send_file(filename, as_attachment=True)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true')
