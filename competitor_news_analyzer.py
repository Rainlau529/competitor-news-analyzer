"""
月报竞品新闻渠道统计分析工具
功能：统计品牌声量和触达数据，按媒体分类输出TOP8排名
"""

import pandas as pd
import os
import sys
from datetime import datetime


# 媒体类别映射表（可扩展）
# 分类依据：结合媒体名称特征、主办方背景、主要内容定位综合判定
MEDIA_CATEGORIES = {
    # ========== 汽车行业垂类媒体 ==========
    # 国内汽车媒体
    "汽车之家": "汽车行业垂类媒体",
    "懂车帝": "汽车行业垂类媒体",
    "易车": "汽车行业垂类媒体",
    "太平洋汽车": "汽车行业垂类媒体",
    "爱卡汽车": "汽车行业垂类媒体",
    "网易汽车": "汽车行业垂类媒体",
    "新浪汽车": "汽车行业垂类媒体",
    "腾讯汽车": "汽车行业垂类媒体",
    "凤凰汽车": "汽车行业垂类媒体",
    "搜狐汽车": "汽车行业垂类媒体",
    "车质网": "汽车行业垂类媒体",
    "中国汽车报": "汽车行业垂类媒体",
    "autohome": "汽车行业垂类媒体",
    "dongchedi": "汽车行业垂类媒体",
    "yiche": "汽车行业垂类媒体",
    # 挪威汽车媒体（名称含"bil"=汽车、"elbil"=电动汽车）
    "Bil24": "汽车行业垂类媒体",          # 挪威汽车新闻
    "BilNorge.no": "汽车行业垂类媒体",    # 挪威汽车新闻
    "Bilnytt.no": "汽车行业垂类媒体",     # 挪威汽车新闻
    "Nybiltester": "汽车行业垂类媒体",    # 新车测试
    "Nyheter fra Bilglede": "汽车行业垂类媒体",  # 汽车新闻
    "Magasinet Motor": "汽车行业垂类媒体",  # 汽车杂志
    "Elbil24.no": "汽车行业垂类媒体",     # 挪威电动汽车新闻（elbil=电动汽车）
    "OtonomHaber (NO)": "汽车行业垂类媒体",  # 自动驾驶/汽车科技新闻（土耳其语"Otonom"=自动驾驶）

    # ========== 商业类媒体 ==========
    # 国内商业媒体
    "第一财经": "商业类媒体",
    "每日经济新闻": "商业类媒体",
    "经济观察报": "商业类媒体",
    "36氪": "商业类媒体",
    "虎嗅": "商业类媒体",
    "钛媒体": "商业类媒体",
    "财经网": "商业类媒体",
    "华尔街见闻": "商业类媒体",
    "界面新闻": "商业类媒体",
    "雪球": "商业类媒体",
    "东方财富": "商业类媒体",
    # 挪威商业/金融媒体
    "E24.no": "商业类媒体",               # 挪威最大商业新闻网站
    "Finansavisen": "商业类媒体",         # 挪威金融新闻
    "Offshore.no": "商业类媒体",          # 挪威海洋/石油/能源行业媒体

    # ========== 技术/科技类媒体 ==========
    # （介于商业与汽车之间，归入商业类更合适）
    "Teknisk Ukeblad Online": "商业类媒体",  # 挪威技术/工程杂志
    "ITavisen.no": "商业类媒体",          # 挪威IT科技新闻
    "Lyd & Bilde Online": "商业类媒体",   # 挪威消费电子评测
    "Tek.no": "商业类媒体",               # 挪威科技新闻
    "Gamereactor (NO)": "商业类媒体",     # 游戏/科技媒体
    "Teksiden": "商业类媒体",             # 挪威技术新闻网站

    # ========== 综合类媒体 ==========
    # 国内综合媒体
    "人民日报": "综合类媒体",
    "新华社": "综合类媒体",
    "央视": "综合类媒体",
    "人民网": "综合类媒体",
    "新华网": "综合类媒体",
    "澎湃新闻": "综合类媒体",
    "今日头条": "综合类媒体",
    "腾讯新闻": "综合类媒体",
    "网易新闻": "综合类媒体",
    "新浪新闻": "综合类媒体",
    "搜狐新闻": "综合类媒体",
    "凤凰资讯": "综合类媒体",
    # 挪威综合媒体（全国性/区域性新闻媒体）
    "TV2 Nyheter": "综合类媒体",           # TV2电视台新闻
    "Dagbladet Online": "综合类媒体",     # 挪威主要报纸
    "VG Online": "综合类媒体",             # 挪威最大报纸
    "Nordlys.no": "综合类媒体",           # 挪威北部地区报纸
    "The News (NO)": "综合类媒体",        # 综合性新闻
    "Norway News (NO)": "综合类媒体",      # 综合新闻
    "Vestlandsnytt": "综合类媒体",        # 挪威西部新闻
    "Dagens (NO)": "综合类媒体",           # 挪威日报
    "MSN": "综合类媒体",                  # 微软MSN新闻聚合

    # ========== 生活类媒体 ==========
    # 国内生活类媒体
    "微信": "生活类媒体",
    "微博": "生活类媒体",
    "小红书": "生活类媒体",
    "抖音": "生活类媒体",
    "快手": "生活类媒体",
    "Bilibili": "生活类媒体",
    "知乎": "生活类媒体",
    "百度": "生活类媒体",
    "UC": "生活类媒体",
    # 挪威生活类媒体
    "Dinside.no": "生活类媒体",           # 挪威消费/生活方式媒体
    "Bodonu": "生活类媒体",               # 生活类网站
}


def format_audience(value):
    """格式化触达人数，添加单位，不四舍五入"""
    if pd.isna(value) or value == 0:
        return "-"

    # 截断到一位小数（不四舍五入）
    if value >= 1_000_000:
        result = value / 1_000_000
        # 截断而非四舍五入
        integer_part = int(result)
        decimal_part = int((result - integer_part) * 10)
        return f"{integer_part}.{decimal_part}M"
    elif value >= 1_000:
        result = value / 1_000
        integer_part = int(result)
        decimal_part = int((result - integer_part) * 10)
        return f"{integer_part}.{decimal_part}K"
    else:
        return str(int(value))


def get_media_category(media_name):
    """获取媒体类别"""
    if pd.isna(media_name):
        return "综合类媒体"

    media_name_str = str(media_name).strip()

    # 精确匹配
    if media_name_str in MEDIA_CATEGORIES:
        return MEDIA_CATEGORIES[media_name_str]

    # 模糊匹配
    for key, category in MEDIA_CATEGORIES.items():
        if key in media_name_str or media_name_str in key:
            return category

    # 默认类别
    return "综合类媒体"


def analyze_competitor_news(input_path):
    """分析竞品新闻渠道数据"""

    print(f"读取文件: {input_path}")
    df = pd.read_excel(input_path)

    # 检查必要的列
    required_columns = ["Media ID", "Date", "Source", "Potential Audience"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        print(f"错误: 缺少必要的列 - {missing_columns}")
        print(f"可用列: {df.columns.tolist()}")
        return None

    print(f"共 {len(df)} 行数据")

    # 转换日期列
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # 获取唯一品牌
    brands = df["Media ID"].dropna().unique().tolist()
    print(f"发现品牌: {brands}")

    # 预计算每个(品牌, 媒体)的最新触达数（避免循环中重复计算）
    latest_audience_cache = {}
    for brand in brands:
        brand_df = df[df["Media ID"] == brand].copy()
        for source in brand_df["Source"].unique():
            media_data = brand_df[brand_df["Source"] == source].copy()
            # 删除日期为空的行
            media_data = media_data.dropna(subset=["Date"])
            if media_data.empty:
                latest_audience_cache[(brand, source)] = None
            else:
                # 获取日期最大的那一行
                latest_idx = media_data["Date"].idxmax()
                latest_row = media_data.loc[latest_idx]
                latest_audience_cache[(brand, source)] = latest_row["Potential Audience"]

    # 存储所有结果
    all_results = []

    for brand in brands:
        print(f"\n处理品牌: {brand}")
        brand_df = df[df["Media ID"] == brand].copy()

        # 统计每个媒体的声量（篇数）
        volume_df = brand_df.groupby("Source").size().reset_index(name="声量（篇）")

        # 获取每个媒体最新日期的触达数（从缓存中取）
        def get_latest_audience(media_name):
            return latest_audience_cache.get((brand, media_name), None)

        volume_df["触达（人次）"] = volume_df["Source"].apply(get_latest_audience)

        # 按声量降序排列，取TOP10
        volume_df = volume_df.sort_values("声量（篇）", ascending=False).head(10)
        volume_df["品牌名称"] = brand
        volume_df["媒体类别"] = volume_df["Source"].apply(get_media_category)

        # 格式化触达人数
        volume_df["触达（人次）"] = volume_df["触达（人次）"].apply(format_audience)

        # 添加序号
        volume_df.insert(0, "序号", range(1, len(volume_df) + 1))

        # 重命名列
        volume_df = volume_df.rename(columns={"Source": "媒体名称"})

        # 重新排列列顺序
        volume_df = volume_df[["序号", "品牌名称", "媒体名称", "声量（篇）", "触达（人次）", "媒体类别"]]

        all_results.append(volume_df)

        print(f"  - {brand}: {len(volume_df)} 条媒体记录")

    # 生成输出文件名
    input_filename = os.path.basename(input_path)
    name_part = os.path.splitext(input_filename)[0]
    output_filename = f"{name_part}处理结果.xlsx"
    output_path = os.path.join(os.path.dirname(input_path), output_filename)

    # 使用openpyxl写入，用空白行分隔品牌
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "处理结果"

    # 表头样式
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_num = 1

    for i, brand_df in enumerate(all_results):
        if i > 0:
            # 添加空白行分隔
            row_num += 1

        # 写入表头
        headers = ["序号", "品牌名称", "媒体名称", "声量（篇）", "触达（人次）", "媒体类别"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row_num += 1

        # 写入数据行
        for _, row_data in brand_df.iterrows():
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            row_num += 1

    # 调整列宽
    column_widths = [8, 12, 25, 12, 14, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    wb.save(output_path)
    print(f"\n结果已保存: {output_path}")

    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python competitor_news_analyzer.py <input.xlsx>")
        print("示例: python competitor_news_analyzer.py 挪威.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]

    if not os.path.exists(input_path):
        print(f"错误: 文件不存在 - {input_path}")
        sys.exit(1)

    result = analyze_competitor_news(input_path)

    if result:
        print(f"\n处理完成!")
